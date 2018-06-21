import xlrd
import copy
import openpyxl
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
import numpy as np
import os
import xlwt
import json
import logging
import sys
import datetime
import time
from pytz import timezone
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from scipy import stats
from itertools import product
import types


BLESSED_TEMP = 22.2222222  # deg C for normalization
CELSIUS_TO_KELVIN = 273.15


class AirPermTest:
    def __init__(self, **kwargs):
        self.specs = list()
        if kwargs:
            specs = []
            for key, value in kwargs.items():
                if key in ['datetime']:
                    setattr(self, key, value)
                elif key == 'baro':
                    self.pressure = value
                elif key[0] == 'p':
                    specs.append({
                        'name': key,
                        'pressures': value
                    })
                    # setattr(self, key, AirPermSpec(pressures=value))
                else:
                    setattr(self, 'temperatures', AirPermThermocouples(**{key: value}))
        else:
            log.error('No input **kwargs for AirPermTest Definition')
            raise TypeError

        for spec in specs:
            self.specs.append(AirPermSpec(name=spec['name'],
                                          time=self.time,
                                          pressures=spec['pressures'],
                                          temperatures=self.avg_temp_kelvin))

    @property
    def avg_temp_celsius(self):
        temps = list()
        if hasattr(self, 'temperatures'):
            for item in vars(self.temperatures):
                values = list()
                for value in getattr(self.temperatures, item):
                    values.append(value)
                temps.append(values)
            return list(np.mean(np.array(temps), axis=0))
        else:
            # given the way that properties seem to work, its okay if this passes through.
            pass

    @property
    def avg_temp_kelvin(self):
        if hasattr(self, 'temperatures'):
            temps = list()
            for item in vars(self.temperatures):
                values = list()
                for value in getattr(self.temperatures, item):
                    values.append(value+CELSIUS_TO_KELVIN)
                temps.append(values)
            return list(np.mean(np.array(temps), axis=0))
        else:
            pass

    @property
    def time(self):
        t0 = time.mktime(self.datetime[0].timetuple())
        t = [time.mktime(tN.timetuple()) - t0 for tN in self.datetime]
        return t


class AirPermSpec:
    def __init__(self, name, time, pressures, temperatures):

        self.name = name
        self.temps = temperatures                               # kelvin
        self.nom_p = pressures                                  # psi
        self.p = self.normalized_pressures(p_array=pressures)   # psi
        self.t = np.array([val/3600 for val in time])
        self.popt, self.pcov = curve_fit(self.exp_model, self.t, self.p)

    def normalized_pressures(self, p_array):
        norm_p = []
        for i, val in enumerate(p_array):
            norm_p.append((val*(BLESSED_TEMP+CELSIUS_TO_KELVIN))/(self.temps[i]))
        return np.array(norm_p)

    def exp_pb(self, conf=0.95):

        xd = self.t
        yd = self.p
        f_vars = self.popt
        x = np.linspace(xd.min(), xd.max(), 100)

        alpha = 1. - conf
        N = xd.size
        variables_n = len(f_vars)

        q = stats.t.ppf(1. - alpha / 2., N - variables_n)

        se = np.sqrt(1. / (N - variables_n) * np.sum((yd - self.exp_model(xd, *f_vars))**2))

        sx = (x - xd.mean()) ** 2
        sxd = np.sum((xd - xd.mean()) ** 2)

        yp = self.exp_model(x, *f_vars)

        dy = q * se * np.sqrt(1. + (1. / N) + (sx / sxd))

        lpb, upb = yp - dy, yp + dy

        return lpb, upb

    def exp_cb(self):
        x = np.linspace(self.t.min(), self.t.max(), 100)

        pos_a = self.popt[0]+(self.pcov[0][0]**0.5)
        neg_a = self.popt[0]-(self.pcov[0][0]**0.5)
        pos_b = self.popt[1]+(self.pcov[1][1]**0.5)
        neg_b = self.popt[1]-(self.pcov[1][1]**0.5)
        pos_c = self.popt[2]+(self.pcov[2][2]**0.5)
        neg_c = self.popt[2]-(self.pcov[2][2]**0.5)
        return self.exp_model(x, pos_a, neg_b, pos_c), self.exp_model(x, neg_a, pos_b, neg_c)

    # def exp_kmpfit(self):
    #     # x = np.linspace(self.t.min(), self.t.max(), 100)
    #     x = self.t
    #     y = self.p
    #     f = kmpfit.simplefit(self.kmpfit_model, [.1, .1, .1], x, y)
    #     a, b, c = f.params
    #     dfdp = [np.exp(-b*x), a*x*np.exp(-b*x), 1]
    #     yhat, upper, lower = f.confidence_band(x, dfdp, 0.95, self.kmpfit_model)
    #
    #     plt.scatter(x, y)
    #     ix = np.argsort(x)
    #     for i, l in enumerate((upper, lower, yhat)):
    #         plt.plot(x[ix], l[ix], c='g' if i == 2 else 'r', lw=2)
    #     plt.show()

    @staticmethod
    def exp_model(t, a, b, c):
        return a * np.exp(-b * t) + c

    @staticmethod
    def kmpfit_model(p, x):
        a, b, c = p
        return a*np.exp(-b*x) + c

    def exp_std_errors(self):
        return np.sqrt(np.diag(self.pcov))

    def plot(self):
        xd = self.t
        yd = self.p
        yd_nom = self.nom_p
        x = np.linspace(xd.min(), xd.max(), 1000)
        popt = self.popt

        lpb, upb = self.exp_pb()
        lcb, ucb = self.exp_cb()

        # plt.plot(x, ucb, c='k', alpha=0.35)
        # plt.plot(x, lcb, c='k', alpha=0.35)
        plt.plot(x, self.exp_model(x, *popt), label=self.name)
        # plt.plot(x, upb, c='r', alpha=0.5)
        # plt.plot(x, lpb, c='r', alpha=0.5)
        # plt.fill_between(x, y1=ucb, y2=lcb, color='k', alpha=0.20, label='95% Confidence Band')
        # plt.fill_between(x, y1=upb, y2=lpb, color='r', alpha=0.20, label='95% Prediction Band')
        plt.scatter(xd, yd, alpha=0.10)
        # plt.scatter(xd, yd_nom, alpha=0.50, c='k', label='Raw Pressures')
        plt.ylabel('Pressure (PSI)')
        plt.xlabel('time (hours)')
        plt.title('Kenda Air Permeation Test: {}'.format(self.name))
        plt.legend()


class AirPermThermocouples:
    def __init__(self, **kwargs):
        if kwargs:
            for key, value in kwargs.items():
                setattr(self, key, value)


def main(input, output='output.xls', template='apt_report_template.xlsx'):
    log.info('CALLING read_xlsx(input_file={})'.format(input))
    apt = read_xlsx(input)
    # log.debug('{}'.format(apt.avg_temp_kelvin[0:5]))
    # log.debug('{}'.format(apt.avg_temp_celsius[0:5]))
    # log.debug('-'*50)
    # log.debug(apt.specs[0].name)
    # log.debug(apt.specs[0].t[20:25])
    # log.debug(apt.specs[0].normalized[20:25])
    log.debug('# of Specs: {}'.format(len(apt.specs)))
    log.debug('-'*50)
    for spec in apt.specs:
        log.debug('Exp. Fit,   STD ERRORS: {}'.format(spec.exp_std_errors()))
        spec.plot()

    plt.show()
    write_xlsx(apt, output, template)


def read_xlsx(file):
    START_ROW = 10

    log.debug('-'*50)
    log.debug('READING INPUT *.XLSX FILE')
    log.debug('-'*50)
    data = {}
    dates = []
    times = []
    wb = xlrd.open_workbook(file)
    log.debug("number of sheets: {}".format(wb.nsheets))
    for i in range(wb.nsheets):
        ws = wb.sheet_by_index(i)
        log.debug("worksheet {}: {}".format(i+1, ws.name))
        # log.debug('-'*50)
        # log.debug('READING COLUMN LABELS, BUILDING DATA DICTIONARY')
        # log.debug('-'*50)
        for j in range(ws.ncols):
            # log.debug("{}".format(ws.col_values(j, 0, 3)))
            if j == 0:
                dates = [datetime.datetime(*xlrd.xldate_as_tuple(val, wb.datemode)) for val in ws.col_values(j, START_ROW)]

            elif j == 1:
                times = [datetime.time((int(val * 24 * 3600)//3600),        # hours
                              (int(val * 24 * 3600) % 3600)//60,   # minutes
                              (int(val * 24 * 3600) % 60))
                         for val in ws.col_values(j, START_ROW)]
            else:
                if isinstance(ws.cell_value(1, j), (int, float)):
                    key = str(ws.cell_value(0, j)) + '_s' + str(int(ws.cell_value(1, j)))
                else:
                    key = str(ws.cell_value(0, j)) + '_' + str(ws.cell_value(1, j))
                key = key.replace('[', '').replace(']', '').replace('°', 'deg_').replace('.', '_').strip('_')
                data[key.lower()] = ws.col_values(j, START_ROW)
                pass

        if dates and times:
            data['datetime'] = [datetime.datetime.combine(date, times[i]) for i, date in enumerate(dates)]
            # log.debug('TOTAL SECONDS OF TIME 0, {}'.format(time.mktime(data['datetime'][0].timetuple())))
    # log.debug('-'*50)
    # log.debug('LOGGING DICTIONARY KEY VALUES')
    # log.debug('-'*50)
    # for key in dict.keys(data):
    #     log.debug(key)
        # if key == 'temp':
            # log.debug("{}".format([key for key in dict.keys(data['temp'])]))
    return AirPermTest(**data)


def write_exp_params(worksheet, spec):
    worksheet['D4'] = spec.popt[0]
    worksheet['F4'] = spec.popt[1]
    worksheet['H4'] = spec.popt[2]
    return worksheet


def write_exp_errors(worksheet, spec):
    std_errors = spec.exp_std_errors()
    worksheet['D5'] = std_errors[0]
    worksheet['F5'] = std_errors[1]
    worksheet['H5'] = std_errors[2]
    return worksheet


def write_temp_constant(worksheet):
    worksheet['E6'] = BLESSED_TEMP + CELSIUS_TO_KELVIN
    return worksheet


def write_spec_name(worksheet, spec):

    worksheet['A1'] = spec.name.upper()
    worksheet.title = spec.name.upper()
    return worksheet


def write_data_arrays(worksheet, spec, n):

    for i in range(n):
        row = 3 + i
        # time
        worksheet.cell(row=row, column=9).value = spec.t[i]
        # nominal pressure
        worksheet.cell(row=row, column=12).value = spec.nom_p[i]
        # averaged temperature
        worksheet.cell(row=row, column=13).value = spec.temps[i]
    return worksheet


def write_formulae(worksheet, n):
    for i in range(n):
        row = 3 + i

        # exp curve values
        time_n = 'I' + str(i+3)  # time (hours)
        worksheet.cell(row=row, column=10).value = '=D4*EXP(-F4*' + time_n + ')+H4'

        # normalized pressure
        pressure_n = 'L'+str(i+3)
        temperature_n = 'M'+str(i+3)
        worksheet.cell(row=row, column=11).value = '=' + pressure_n + '* E6/' + temperature_n

        # lower confidence band
        worksheet.cell(row=row, column=14).value = '=(D4-D5)*EXP(-(F4-F5)*'+time_n+')+(H4-H5)'

        # upper confidence band
        worksheet.cell(row=row, column=15).value = '=(D4+D5)*EXP(-(F4+F5)*'+time_n+')+(H4+H5)'

    return worksheet


def write_plots(worksheet, spec):
    chart = ScatterChart()
    chart.title = 'Measured Data w/ Exponential Fit \n Pressure vs. Temperature'
    chart.style = 13
    chart.x_axis.title = 'Time (hrs)'
    chart.y_axis.title = 'Pressure (psi)'

    # time values (hours)
    x_values = Reference(worksheet, min_col=9, min_row=3, max_row=len(spec.p) + 3)

    # pressure data (psi)
    p = Reference(worksheet, min_col=11, min_row=3, max_row=len(spec.p) + 3)
    p_series = Series(p, x_values)
    chart.series.append(p_series)

    # exponential curve fit data (psi)
    exp = Reference(worksheet, min_col=10, min_row=3, max_row=len(spec.p) + 3)
    exp_series = Series(exp, x_values)
    chart.series.append(exp_series)

    # Style the lines
    s1 = chart.series[1]
    s1.marker.symbol = "triangle"
    s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline

    s1.graphicalProperties.line.noFill = True

    s2 = chart.series[0]
    s2.graphicalProperties.line.solidFill = "00AAAA"
    s2.graphicalProperties.line.dashStyle = "sysDot"
    s2.graphicalProperties.line.width = 100050  # width in EMUs

    worksheet.add_chart(chart, 'A7')
    return worksheet


def write_styling(worksheet):
    return worksheet


def write_spec_summary(wb, i, worksheet, spec):
    summary_label_offset = 7
    summary_label_gap = 3

    # row number that data should be located at, function of spec # (aka worksheet #)
    # n = first spec row
    # n1 = second spec row
    n = str((summary_label_gap * i) + summary_label_offset)
    n1 = str((summary_label_gap * i) + summary_label_offset + 1)
    wb['SUMMARY']['B' + n] = spec.name.upper()
    wb['SUMMARY']['E' + n] = '='+spec.name + '!K3'
    wb['SUMMARY']['E' + n1] = '='+spec.name + '!J3'

    # spec_name needed
    exp_curve_eval = '=' + spec.name + '!D4*EXP(-'+spec.name + \
                     '!F4*((720*SUMMARY!G5)+(24*SUMMARY!H5)+SUMMARY!I5))+' + spec.name + '!H4'

    lin_data_interp = '=FORECAST(((720*G5)+(24*H5)+(I5)),' + \
                      'OFFSET('+spec.name+'!K3:K590,MATCH(((720*G5)+(24*H5)+(I5)),'+spec.name+'!I3:I590,1)-1,0,2),' + \
                      'OFFSET('+spec.name+'!I3:I590,MATCH(((720*G5)+(24*H5)+(I5)),'+spec.name+'!I3:I590,1)-1,0,2))'

    wb['SUMMARY']['G'+n] = lin_data_interp
    wb['SUMMARY']['G'+n1] = exp_curve_eval
    return wb


def write_xlsx(data, file, template):

    wb = openpyxl.load_workbook(template)

    for i, spec in enumerate(data.specs):
        target = 'SPEC_'+str(i+1)
        ws = wb[target]

        ws = write_exp_params(ws, spec)
        ws = write_exp_errors(ws, spec)
        ws = write_temp_constant(ws)
        ws = write_spec_name(ws, spec)
        ws = write_data_arrays(ws, spec, len(spec.p))
        ws = write_formulae(ws, len(spec.p))
        ws = write_plots(ws, spec)
        ws = write_styling(ws)
        wb = write_spec_summary(wb, i, ws, spec)

    # ws1 = wb['SUMMARY']

    wb.save('output.xlsx')


def init_logger():
    filelog = logging.getLogger('status')
    filelog.setLevel(logging.DEBUG)

    formatter = logging.Formatter('%(name)s - %(funcName)15s - %(lineno)d - %(levelname)s - %(message)s')

    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)

    fh = logging.FileHandler(filename='kpat_results.log', mode='a')
    fh.setLevel(logging.DEBUG)

    ch.setFormatter(formatter)
    fh.setFormatter(formatter)

    filelog.addHandler(fh)
    filelog.addHandler(ch)
    return filelog


if __name__ == "__main__":

    fmt = '%Y-%m-%d | %H:%M:%S'

    log = init_logger()
    log.info('==============================================================================')
    log.info('----------------          STARTING-UP THIS BIATCH         --------------------')
    log.info('-------------  BICYCLE GROUP AIR-PERM TEST, POST PROCESSING  -----------------')
    log.info('----------------           {}          --------------------'.format(
        str(datetime.datetime.now(timezone('US/Eastern')).strftime(fmt))
    ))
    log.info('==============================================================================')

    if len(sys.argv) >= 3:
        main(input=sys.argv[1], output=sys.argv[2])

    elif len(sys.argv) == 2:
        main(input=sys.argv[1])

    else:
        log.error("INCOMPETENCE ERROR: No input file, SOLUTION: I suggest to stop being stupid")
        log.info('PROGRAM TERMINATED WITH ERRORS - USER IS BEING A MORON')

    log.info('==============================================================================')
    log.info('----------------            {}         --------------------'.format(
        str(datetime.datetime.now(timezone('US/Eastern')).strftime(fmt))
    ))
    log.info('----------------                    DONE                  --------------------')
    log.info('-------------  BICYCLE GROUP AIR-PERM TEST, POST PROCESSING  -----------------')
    log.info('----------------                 PROGRAM END              --------------------')
    log.info('==============================================================================')